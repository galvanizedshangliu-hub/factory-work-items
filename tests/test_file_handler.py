"""
本地测试 file_handler 的文件解析功能
不需要钉钉网络连接，只测 parse_spreadsheet + format_file_summary
"""
import sys, os, io
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.file_handler import parse_spreadsheet, format_file_summary


def test_excel():
    """用 openpyxl 生成测试 Excel 并解析"""
    from openpyxl import Workbook

    # 创建测试 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "订单需求"
    ws.append(["订单号", "客户", "规格", "数量(件)", "交期", "优先级", "备注"])
    ws.append(["ORD001", "客户A", "Φ15.2", "500", "2026-06-20", "高", "急单"])
    ws.append(["ORD002", "客户B", "Φ12.7", "300", "2026-06-22", "中", ""])
    ws.append(["ORD003", "客户C", "Φ15.2", "800", "2026-06-18", "高", ""])

    # 第二个工作表
    ws2 = wb.create_sheet("库存")
    ws2.append(["规格", "库存量", "安全库存"])
    ws2.append(["Φ15.2", "300", "100"])
    ws2.append(["Φ12.7", "500", "80"])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    file_bytes = buf.getvalue()

    print("=" * 60)
    print("测试 1: Excel 解析 (.xlsx)")
    print("=" * 60)
    result = parse_spreadsheet(file_bytes, "test_order.xlsx")
    summary = format_file_summary(result, "test_order.xlsx")
    print(summary)
    print()
    assert result["success"], "解析失败"
    assert result["row_count"] == 7, f"总行数不对: {result['row_count']}"
    assert len(result["sheet_names"]) == 2, f"工作表数不对: {len(result['sheet_names'])}"
    assert "库存" in result["sheet_names"]
    print("✅ Excel 解析测试通过！")
    print()


def test_csv():
    """生成测试 CSV 并解析"""
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["订单号", "客户", "规格", "数量", "交期"])
    writer.writerow(["ORD-A01", "甲公司", "Φ15.2", "1200", "2026/6/25"])
    writer.writerow(["ORD-A02", "乙公司", "Φ17.8", "600", "2026/6/28"])
    csv_content = buf.getvalue()

    # 同时测试带 BOM 的 UTF-8
    file_bytes = csv_content.encode("utf-8-sig")

    print("=" * 60)
    print("测试 2: CSV 解析 (.csv)")
    print("=" * 60)
    result = parse_spreadsheet(file_bytes, "order_csv.csv")
    summary = format_file_summary(result, "order_csv.csv")
    print(summary)
    print()
    assert result["success"], "解析失败"
    assert result["row_count"] == 3, f"总行数不对: {result['row_count']}"
    print("✅ CSV 解析测试通过！")
    print()


def test_unsupported_format():
    """测试不支持的文件格式"""
    print("=" * 60)
    print("测试 3: 不支持的文件格式")
    print("=" * 60)
    result = parse_spreadsheet(b"not used", "file.pdf")
    assert not result["success"], "应该失败"
    assert "不支持" in result["error"]
    print(f"正确拦截: {result['error']}")
    print("✅ 格式拦截测试通过！")
    print()


def test_empty_excel():
    """测试空 Excel"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "空表"
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()

    print("=" * 60)
    print("测试 4: 空 Excel")
    print("=" * 60)
    result = parse_spreadsheet(buf.getvalue(), "empty.xlsx")
    summary = format_file_summary(result, "empty.xlsx")
    print(summary)
    assert result["success"]
    assert result["row_count"] == 0
    print("✅ 空 Excel 测试通过！")
    print()


if __name__ == "__main__":
    test_excel()
    test_csv()
    test_unsupported_format()
    test_empty_excel()
    print("🎉 全部测试通过！")
